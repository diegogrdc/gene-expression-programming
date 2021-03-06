from Chromosome import Chromosome
from GeneExpressor import GeneExpressor
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime
from os.path import exists
import os
import re
import pickle


class FitnessEvaluator:
    # We initialize the fitness evaluator
    # - dir_cases = Directory Where we will find our fitness cases
    # - dir_test_cases = Directory Where we will find our test cases
    # - function_set = Functions that will be used for expression
    # Given by the user
    def __init__(self, dir_cases, dir_test_cases, function_set):
        self.training_dir = dir_cases
        self.test_dir = dir_test_cases
        # We extract all the cases from the directory, and store them
        self.initTrainingCases()
        self.initTestCases()
        # Set up gene expression class to deal with
        # gene expression when needed
        # We store function set there
        self.gene_expressor = GeneExpressor(function_set)

        # We setup our heuristics
        # This is problem specific
        # In this case we have four
        # TODO UNCOMMENT
        self.heuristics = [self.heuristic_next_fit, self.heuristic_first_fit,
                           self.heuristic_best_fit, self.heuristic_worst_fit]
        # Setup excel stuff
        self.setUpExcel()

    # Open excel worsheet to setup data after we finish testing, to store
    # results for simple heuristics, and then update with our best tests as needed
    def setUpExcel(self):
        if exists('ResHWF.xlsx'):
            self.workbook = load_workbook('ResHWF.xlsx')
            self.worksheet = self.workbook['Results']
        else:
            self.workbook = Workbook()
            self.workbook.create_sheet("Results")
            self.worksheet = self.workbook['Results']
            # If first time, setup simple heuristics and names
            # Set Titles
            self.worksheet.cell(2, 1, "Test Names")
            self.worksheet.cell(2, 2, "Next Fit")
            self.worksheet.cell(2, 3, "First Fit")
            self.worksheet.cell(2, 4, "Best Fit")
            self.worksheet.cell(2, 5, "Worst Fit")
            for i in range(0, len(self.test_names)):
                # Set test names in col 1
                self.worksheet.cell(3 + i, 1, self.test_names[i])
                # Set heuristic 0 in col 2
                res = 1000 - self.evalFitnessCase([], i, False, 0)
                self.worksheet.cell(3 + i, 2, res)
                # Set heuristic 1 in col 3
                res = 1000 - self.evalFitnessCase([], i, False, 1)
                self.worksheet.cell(3 + i, 3, res)
                # Set heuristic 2 in col 4
                res = 1000 - self.evalFitnessCase([], i, False, 2)
                self.worksheet.cell(3 + i, 4, res)
                # Set heuristic 3 in col 5
                res = 1000 - self.evalFitnessCase([], i, False, 3)
                self.worksheet.cell(3 + i, 5, res)
            self.workbook.save('ResHWF.xlsx')

    def heuristic_next_fit(self, terminals):
        # Get last bin
        # + 4 to adjust as we have 4 nums unrelated before
        # - 1 to index in zero
        bin = terminals[2] - 1 + 4
        c = terminals[1]
        nw = terminals[3]
        # If new item fits we just add it in same bin
        if terminals[bin] + nw <= c:
            terminals[bin] = terminals[bin] + nw
        # Else we create new bin and add it there
        else:
            terminals[2] = terminals[2] + 1
            terminals[bin + 1] = nw

    # Consider all bins, and  find first fit
    # If did not find any, add a new bin and insert
    def heuristic_first_fit(self, terminals):
        c = terminals[1]
        nw = terminals[3]
        bin = terminals[2] - 1 + 4
        # Search bins left to right
        for i in range(4, len(terminals)):
            # If it fits, we store and return
            if terminals[i] + nw <= c:
                terminals[i] = terminals[i] + nw
                # Check if new bin
                if i > bin:
                    terminals[2] = terminals[2] + 1
                return
        # We should never arrive here, as we should always find a place to store
        assert(0 == 1)

    # Consider all bins, and  find best fit
    # best means one with least residue left
    # If did not find any, add a new bin and insert
    def heuristic_best_fit(self, terminals):
        c = terminals[1]
        nw = terminals[3]
        bin = terminals[2] - 1 + 4
        bst = -1
        bst_fit = c + 1
        for i in range(4, len(terminals)):
            if terminals[i] + nw <= c and c - nw - terminals[i] < bst_fit:
                bst_fit = c - nw - terminals[i]
                bst = i
        if bst > bin:
            terminals[2] = terminals[2] + 1
        terminals[bst] = terminals[bst] + nw

    # Consider all bins, and  find worst fit
    # worst means one with most residue left
    # If did not find any, add a new bin and insert
    def heuristic_worst_fit(self, terminals):
        c = terminals[1]
        nw = terminals[3]
        bin = terminals[2] - 1 + 4
        wst = -1
        wst_fit = -1
        for i in range(4, bin + 1):
            if terminals[i] + nw <= c and c - nw - terminals[i] > wst_fit:
                wst_fit = c - nw - terminals[i]
                wst = i
        if wst == -1:
            terminals[2] = terminals[2] + 1
            wst = bin + 1
        terminals[wst] = terminals[wst] + nw

    # With directory name, we store all our fitness cases
    # NOTE: This is specific to the problem, this should change
    # in a different problem if needed
    def initTestCases(self):
        # This will contain test names
        self.test_names = []
        # This will contain test size n, number of objects
        self.test_n = []
        # This will contain test val c, capacity of bins
        self.test_c = []
        # This will contain test object values for each test
        self.test_objs = []
        # Iterate over every file in our directory
        for filename in os.listdir(self.test_dir):
            # We just check bpp test files
            if not "bpp" in filename:
                continue
            # Store test name
            self.test_names.append(filename)
            # Get test contents
            file = open(self.test_dir + filename, 'r')
            content = file.read()
            # Parse contents and store them in list as ints
            nums = re.findall(r'\d+', content)
            nums = list(map(int, nums))
            # Store relevant information
            self.test_n.append(nums[0])
            self.test_c.append(nums[1])
            self.test_objs.append(nums[2:])
            # Make sure tests are ok, n == total objs
            assert(nums[0] == len(nums[2:]))

    # With directory name, we store all our fitness cases
    # NOTE: This is specific to the problem, this should change
    # in a different problem if needed
    def initTrainingCases(self):
        # This will contain test names
        self.train_names = []
        # This will contain test size n, number of objects
        self.train_n = []
        # This will contain test val c, capacity of bins
        self.train_c = []
        # This will contain test object values for each test
        self.train_objs = []
        # Iterate over every file in our directory
        for filename in os.listdir(self.training_dir):
            # We just check bpp test files
            if not "bpp" in filename:
                continue
            # Store test name
            self.train_names.append(filename)
            # Get test contents
            file = open(self.training_dir + filename, 'r')
            content = file.read()
            # Parse contents and store them in list as ints
            nums = re.findall(r'\d+', content)
            nums = list(map(int, nums))
            # Store relevant information
            self.train_n.append(nums[0])
            self.train_c.append(nums[1])
            self.train_objs.append(nums[2:])
            # Make sure tests are ok, n == total objs
            assert(nums[0] == len(nums[2:]))

    # Function to evaluate the fitness of a specific population
    # - population = List with Population, each element representing a
    # - trainf = Flag used to determine test or train set
    # chromosome generated in process
    # Return values obtained in a list
    def evaluatePopulationFitness(self, population, trainf):
        fitness = []
        for ind in population:
            fitness.append(
                1000 - self.evaluateChromosomeFitness(ind, 0, trainf, None))
        print("Avg fitness are", fitness)
        print("Best fitness is", max(fitness))
        self.storePopulation(population)
        return fitness

    # Using pickle to store population and not start
    # from 0 every time
    def storePopulation(self, population):
        filename = 'best_store.txt'
        outfile = open(filename, 'wb')
        pickle.dump(population, outfile)
        outfile.close()

    # Evaluates fitness of an individual, by evaluating each one of our
    # test cases, and getting and average of evaluations
    # - ind = Individual chromosome used to evaluate fitness
    # - f = Flag used for printing if necessary
    # - trainf = Flag used to determine test or train set
    def evaluateChromosomeFitness(self, ind, f, trainf, heuristic):
        waste = []
        # For each test, we get fitness
        names = self.train_names if trainf else self.test_names
        for i in range(0, len(names)):
            waste.append(self.evalFitnessCase(ind, i, trainf, heuristic))
        avgwaste = sum(waste) / len(waste)
        if trainf:
            return avgwaste
        return waste

    def writeTestResult(self, ind, waste, col):
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
        self.worksheet.cell(2, col, "GEP Exp" + dt_string)
        for i in range(0, len(self.test_names)):
            self.worksheet.cell(3 + i, col, 1000 - waste[i])
        self.worksheet.cell(1, col, str(ind))
        self.workbook.save('ResHWF.xlsx')

    # Evaluate a chromosome in a specific test
    # - ind = Individual chromosome used to evaluate fitness
    # - idx = Index of test used
    # - trainf = Flag used to determine test or train set
    # This function is also specific to the problem in hand
    # It needs for test cases to all have the same size
    # In this case we have 20 objects
    # If we change, it does not work
    def evalFitnessCase(self, ind, idx, trainf, heuristic):
        # Get our data (test or train)
        n = self.train_n if trainf else self.test_n
        c = self.train_c if trainf else self.test_c
        objs = self.train_objs if trainf else self.test_objs
        # We setup our "environment"
        # This is the terminals that the problem will use for its expression
        # We have a list of length 24
        # First element is N, dynamic on each object
        # Second element is C
        # Third element is current # of bins used
        # Fourth element is arriving object weigth
        terminals = [n[idx], c[idx], 1, 0]
        # Next 20 elements are the N possible bins we can have at the end
        # each with its current occupancy
        terminals.extend([0] * n[idx])
        chosens = [0, 0, 0, 0]
        # Now we simulate each items arrival
        for i, item in enumerate(objs[idx]):
            # Update terminals with given object
            if item > c[idx]:
                # Item is bigger than capacity
                continue
            terminals[0] = (i + 1) / n[idx]
            terminals[3] = item

            # Pick heuristic with gene expression and with given terminals
            if heuristic == None:
                heuristic = self.chooseHeuristic(ind, terminals)

            # Check which heuristic to use and do it
            self.heuristics[heuristic](terminals)
            chosens[heuristic] = chosens[heuristic] + 1

        diff_chosens = 0
        for i in range(0, 4):
            if chosens[i] > 0:
                diff_chosens = diff_chosens + 1
        if diff_chosens > 1:
            print("chosens: ", chosens)
        # After all simulation, we calculate fitness with our formula
        # We get all wasted space
        waste = 0
        bin = terminals[2] + 4
        C = terminals[1]
        for i in range(4, bin):
            waste = waste + C - terminals[i]

        return waste

    # This function calls the gene expressor
    # to express our gene and get the chosen
    # heuristic to follow up the problem
    # - ind = Gene we will express
    # with its adfs and homeotic genes
    # - terminals = Set of terminals
    # used in this state of problem
    def chooseHeuristic(self, ind, terminals):
        # Get a prediction val for each class
        heuristics = self.gene_expressor.express(ind, terminals)
        # Return best heuristic
        chosen = np.argmax(heuristics)
        return chosen
